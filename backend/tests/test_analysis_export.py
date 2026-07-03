"""X5 — the export engine (registry: CSV + XLSX + repro-card + LaTeX) + its API routes.

Covers the Definition of Done: a tidy full-resolution CSV that pandas can read; an XLSX with the right
sheets, a known cell, and a native chart; both pivots (per-game raw / per-algorithm normalized); the
"complete workbook" (a sheet per game with ≥1 history); a deterministic config-hash; a valid BibTeX +
reproduce command; a booktabs LaTeX table; and the LTTB helper preserving a curve's shape + endpoints.
"""

import csv
import io
from pathlib import Path

import openpyxl
from app.api import analysis as analysis_api
from app.main import app
from app.schemas.training import TrainConfig
from app.services.analysis import export as export_engine
from app.services.analysis import provenance
from app.services.analysis.lttb import downsample
from app.services.runs import RunStore
from fastapi.testclient import TestClient


def _frame(env_steps: int, ep_rew_mean: float | None, elapsed: float, **extra: object) -> dict:
    return {
        "type": "metrics", "env_steps": env_steps, "timesteps": env_steps,
        "ep_rew_mean": ep_rew_mean, "wall_clock": elapsed, "elapsed": elapsed, **extra,
    }


def _make_store(tmp_path: Path, monkeypatch) -> RunStore:
    """A RunStore pointed at a tmp dir, injected into both the export engine and the API module."""
    store = RunStore(tmp_path / "runs")
    monkeypatch.setattr(export_engine, "run_store", store)
    monkeypatch.setattr(analysis_api, "run_store", store)
    return store


def _cartpole_ppo(store: RunStore) -> str:
    cfg = TrainConfig(env_id="cartpole", algo="ppo", seed=7)
    frames = [_frame(2048, 30.0, 1.0, ep_len_mean=30.0, loss=0.5),
              _frame(4096, 300.0, 2.0), _frame(6144, 500.0, 3.0)]
    return store.save(cfg, frames, state="finished",
                      started_at="2026-07-02T10:00:00+00:00", solved_score=500.0).id


def _cartpole_evo(store: RunStore) -> str:
    cfg = TrainConfig(env_id="cartpole", algo="neuroevolution", seed=1,
                      evolution={"population_size": 40})
    frames = [
        {"type": "evolution", "env_steps": 6000, "timesteps": 6000, "best_fitness": 120.0,
         "wall_clock": 4.0, "elapsed": 4.0},
        {"type": "evolution", "env_steps": 12000, "timesteps": 12000, "best_fitness": 480.0,
         "wall_clock": 8.0, "elapsed": 8.0},
    ]
    return store.save(cfg, frames, state="finished",
                      started_at="2026-07-02T11:00:00+00:00", solved_score=500.0).id


# -- LTTB helper -----------------------------------------------------------


def test_lttb_keeps_endpoints_and_caps_length() -> None:
    xs = list(range(1000))
    ys = [float(x) for x in xs]
    out = downsample(xs, ys, 100)
    assert len(out) == 100
    assert out[0] == (0.0, 0.0) and out[-1] == (999.0, 999.0)  # endpoints preserved


def test_lttb_passthrough_when_small_and_drops_none() -> None:
    xs = [0, 1, 2]
    ys = [0.0, None, 2.0]
    out = downsample(xs, ys, 800)  # fewer than threshold → returned as-is, None dropped
    assert out == [(0.0, 0.0), (2.0, 2.0)]


def test_lttb_preserves_a_spike() -> None:
    # A flat line with one tall spike — LTTB must keep the spike, a naive stride could miss it.
    xs = list(range(200))
    ys = [0.0] * 200
    ys[97] = 100.0
    out = downsample(xs, ys, 20)
    assert any(y == 100.0 for _, y in out)


# -- CSV (tidy/long) -------------------------------------------------------


def test_csv_is_tidy_long_full_resolution(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    rid = _cartpole_ppo(store)
    content, media, fname = export_engine.export("csv", [rid])
    assert media == "text/csv" and fname.endswith(".csv")
    rows = list(csv.DictReader(io.StringIO(content.decode("utf-8"))))
    assert rows[0].keys() == {  # exact tidy schema
        "run_id", "experiment_id", "env_id", "algo", "seed",
        "env_steps", "wall_clock", "metric", "value",
    }
    # Full resolution: every frame's reward is present (3 frames → 3 reward rows).
    reward_rows = [r for r in rows if r["metric"] == "reward"]
    assert len(reward_rows) == 3
    assert [r["value"] for r in reward_rows] == ["30.0", "300.0", "500.0"]
    # Both pivots derivable: skill_pct rows exist too (30/500 = 6%).
    skill_rows = [r for r in rows if r["metric"] == "skill_pct"]
    assert len(skill_rows) == 3 and skill_rows[0]["value"] == "6.0"
    # Extra per-frame metrics only where populated (loss/ep_len only on the first frame here).
    assert {r["metric"] for r in rows if r["metric"] in ("loss", "ep_len_mean")} == {"loss", "ep_len_mean"}


# -- XLSX ------------------------------------------------------------------


def test_xlsx_has_expected_sheets_and_known_cell(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    p_id = _cartpole_ppo(store)
    e_id = _cartpole_evo(store)
    content, media, fname = export_engine.export("xlsx", [p_id, e_id])
    assert media.endswith("spreadsheetml.sheet") and fname.endswith(".xlsx")

    wb = openpyxl.load_workbook(io.BytesIO(content))
    # Summary + one per-game sheet (both runs are cartpole) + Config + Methods.
    assert "Summary" in wb.sheetnames and "Config" in wb.sheetnames and "Methods" in wb.sheetnames
    game_sheets = [s for s in wb.sheetnames if s not in ("Summary", "Config", "Methods")]
    assert len(game_sheets) == 1  # complete-workbook: one sheet for the one game with history

    summary = wb["Summary"]
    assert summary["A3"].value == "run_id"  # header row
    # A known cell: the first data row's run_id is one of the two saved runs.
    assert summary["A4"].value in {p_id, e_id}

    # The per-game sheet carries a native chart.
    assert len(wb[game_sheets[0]]._charts) == 1


def test_xlsx_complete_workbook_one_sheet_per_game(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    cart = _cartpole_ppo(store)
    # A second game.
    cfg = TrainConfig(env_id="acrobot", algo="ppo", seed=3)
    acro = store.save(cfg, [_frame(1000, -400.0, 1.0), _frame(2000, -110.0, 2.0)],
                      state="finished", started_at="2026-07-02T12:00:00+00:00",
                      solved_score=-100.0).id
    content, _, _ = export_engine.export("xlsx", [cart, acro])
    wb = openpyxl.load_workbook(io.BytesIO(content))
    game_sheets = [s for s in wb.sheetnames if s not in ("Summary", "Config", "Methods")]
    assert len(game_sheets) == 2  # a sheet for each game with ≥1 history


def test_xlsx_algo_pivot_uses_normalized_skill(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    p_id = _cartpole_ppo(store)
    content, _, _ = export_engine.export("xlsx", [p_id], pivot="algo")
    wb = openpyxl.load_workbook(io.BytesIO(content))
    # Per-algorithm pivot → the group sheet is named for the algo, and its curve is skill-% (0–100),
    # not raw reward: the final point normalizes 500 reward → 100 %.
    ppo_sheet = wb["ppo"]
    header_cells = [c.value for row in ppo_sheet.iter_rows() for c in row if c.value]
    assert any("skill %" in str(v) for v in header_cells)


# -- Reproducibility card + config-hash determinism ------------------------


def test_config_hash_is_deterministic_and_field_sensitive() -> None:
    a = TrainConfig(env_id="cartpole", algo="ppo", seed=7)
    b = TrainConfig(env_id="cartpole", algo="ppo", seed=7)
    c = TrainConfig(env_id="cartpole", algo="ppo", seed=8)
    h = provenance.config_hash(a)
    assert h == provenance.config_hash(b)  # same config → same hash
    assert len(h) == 64 and all(ch in "0123456789abcdef" for ch in h)  # sha256 hex
    assert provenance.config_hash(c) != h  # a different seed changes it


def test_repro_card_has_hash_bibtex_and_command(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    rid = _cartpole_ppo(store)
    content, media, fname = export_engine.export("repro", [rid])
    assert media == "text/markdown" and fname.endswith(".md")
    text = content.decode("utf-8")
    assert provenance.config_hash(TrainConfig(env_id="cartpole", algo="ppo", seed=7)) in text
    assert "@software{" in text  # BibTeX entry
    assert "curl -X POST" in text and "/api/train/start" in text  # reproduce command


def test_bibtex_year_comes_from_run_not_export() -> None:
    cfg = TrainConfig(env_id="cartpole", algo="ppo", seed=7)
    bib = provenance.bibtex(cfg, created_at="2024-01-15T10:00:00+00:00", label="my run")
    assert "year   = {2024}" in bib  # the run's year, deterministic


# -- LaTeX -----------------------------------------------------------------


def test_latex_is_booktabs_table(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    rid = _cartpole_ppo(store)
    content, media, fname = export_engine.export("latex", [rid])
    assert fname.endswith(".tex")
    text = content.decode("utf-8")
    assert "\\toprule" in text and "\\midrule" in text and "\\bottomrule" in text
    assert "\\begin{tabular}" in text
    assert "cartpole" in text  # a data row


# -- API routes ------------------------------------------------------------


def test_export_routes_return_downloads(tmp_path: Path, monkeypatch) -> None:
    store = _make_store(tmp_path, monkeypatch)
    rid = _cartpole_ppo(store)
    client = TestClient(app)
    for path, ctype in (
        ("/api/analysis/export.csv", "text/csv"),
        ("/api/analysis/export.repro", "text/markdown"),
        ("/api/analysis/export.tex", "text/plain"),
    ):
        resp = client.get(path, params={"run_ids": [rid]})
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(ctype)
        assert "attachment" in resp.headers["content-disposition"]
    # XLSX is binary and opens with openpyxl.
    resp = client.get("/api/analysis/export.xlsx", params={"run_ids": [rid]})
    assert resp.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert "Summary" in wb.sheetnames


def test_export_empty_selection_is_valid_not_error() -> None:
    client = TestClient(app)
    resp = client.get("/api/analysis/export.csv")  # no run_ids
    assert resp.status_code == 200
    # A valid header-only CSV (the client can still hand the user a file).
    assert resp.content.decode("utf-8").startswith("run_id,experiment_id")


def test_export_unknown_run_id_skipped(tmp_path: Path, monkeypatch) -> None:
    _make_store(tmp_path, monkeypatch)
    client = TestClient(app)
    resp = client.get("/api/analysis/export.csv", params={"run_ids": ["does-not-exist"]})
    assert resp.status_code == 200
    rows = list(csv.DictReader(io.StringIO(resp.content.decode("utf-8"))))
    assert rows == []  # header only, no data — unknown id dropped, not fatal
