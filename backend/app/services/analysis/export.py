"""Export engine (X5, Wave 1) — turn a selection of runs into a citable dataset.

The headline DataLab deliverable: one pipeline ``load run(s) → normalized frames → format``, where each
output format is a thin **plugin** over the shared loading + normalization, so a new format (the Wave-3
TensorBoard / vector-figure plugins) slots in without touching the loaders. Everything is server-side —
the frontend store is a capped live ring buffer, so the *full* raw history on disk (``data/runs/``) is the
only honest export source.

Wave 1 ships four zero-heavy-dep formats over the registry:

* **CSV (tidy/long)** — one row per ``(run, frame, metric)`` at **full resolution** (pandas/R). Carries
  both the raw ``reward`` and the normalized ``skill_pct`` metric, so *both* pivots (per-game raw /
  per-algorithm normalized) are a ``groupby`` away.
* **XLSX (publication layout)** — ``Summary`` (the X2 stats, one row per run) + a per-game sheet for every
  game with ≥1 run (config header + an LTTB-downsampled curve table + a **native Excel chart**) + ``Config``
  (full hyperparameters) + ``Methods`` (versions / hardware / git + the reproducibility card). The ``pivot``
  picks the per-game sheet's metric: ``"game"`` = raw reward, ``"algo"`` = normalized skill-%.
* **Reproducibility card** — the citable glue (config-hash + BibTeX + reproduce command), as its own
  Markdown artifact and embedded in the XLSX ``Methods`` sheet. Built in :mod:`.provenance`.
* **LaTeX table** — the X2 summary stats as a paste-ready booktabs table (pure text, no dep).

Column *headers on the data tables are canonical English identifiers* (they load straight into
pandas/R); the *descriptive* labels a human reads — the Methods sheet fields, the per-game config header,
the sheet section titles — are bilingual (EN + CZ), since the backend has no i18n runtime.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from dataclasses import dataclass
from typing import Any, Literal

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.marker import Marker
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.envs.registry import get_env
from app.schemas.analysis import RunSummary
from app.schemas.runs import RunMeta
from app.schemas.training import TrainConfig
from app.services.analysis import provenance
from app.services.analysis import rliable_metrics as rliable
from app.services.analysis.lttb import downsample
from app.services.analysis.stats import score_of_frame, skill_pct, summarize
from app.services.runs import run_store

Pivot = Literal["game", "algo"]

# Target points per curve in the XLSX sheets/charts (CSV stays full-resolution). ~800 keeps the shape
# visually lossless while staying far under Excel's row cap, so curves are never silently truncated.
_LTTB_TARGET = 800

# The per-run config attribute that carries the active algorithm's tunable hyperparameters.
_ALGO_BLOCK: dict[str, str] = {
    "ppo": "hyperparams", "neuroevolution": "evolution", "q_learning": "q_learning",
    "alphazero": "alphazero", "sac": "sac", "td3": "td3", "dqn": "dqn",
}


@dataclass
class LoadedRun:
    """One run loaded + normalized for export: its provenance, config, raw frames, X2 summary, and the
    env's display names + skill range (looked up once so the formatters need not re-hit the registry)."""

    meta: RunMeta
    config: TrainConfig
    frames: list[dict[str, Any]]
    summary: RunSummary
    env_name_en: str
    env_name_cz: str
    min_score: float
    solved_score: float


def load_runs(run_ids: list[str]) -> list[LoadedRun]:
    """Load + normalize each requested run from disk, skipping unknown ids (one bundle per *found* run).

    Mirrors ``GET /api/analysis/summary``: the store backfills the X1 canonical axes on read, and each
    curve is normalized against its env's ``[min_score, solved_score]``. Unreadable / missing runs are
    dropped rather than failing the whole export.
    """
    loaded: list[LoadedRun] = []
    for rid in run_ids:
        detail = run_store.get(rid)
        if detail is None:
            continue
        spec = get_env(detail.config.env_id)
        min_score = spec.min_score if spec else 0.0
        solved_score = spec.solved_score if spec else 0.0
        summary = summarize(
            run_id=detail.meta.id, env_id=detail.config.env_id, algo=detail.config.algo,
            seed=detail.config.seed, frames=detail.metrics,
            min_score=min_score, solved_score=solved_score,
        )
        loaded.append(
            LoadedRun(
                meta=detail.meta, config=detail.config, frames=detail.metrics, summary=summary,
                env_name_en=spec.display_name.en if spec else detail.config.env_id,
                env_name_cz=spec.display_name.cz if spec else detail.config.env_id,
                min_score=min_score, solved_score=solved_score,
            )
        )
    return loaded


def active_hyperparams(config: TrainConfig) -> dict[str, Any]:
    """The tunable hyperparameters of the run's *active* algorithm (the block ``algo`` selects), flattened
    to ``{name: value}``. Empty if the block is absent (a legacy config that only carried PPO defaults)."""
    block = getattr(config, _ALGO_BLOCK.get(config.algo, ""), None)
    return block.model_dump() if block is not None else {}


# ---------------------------------------------------------------------------
# CSV (tidy / long) — full resolution
# ---------------------------------------------------------------------------

# The metrics emitted per frame, each as its own tidy row. Both pivots are derivable downstream:
# ``reward`` (raw, the per-game view) and ``skill_pct`` (normalized, the per-algorithm view).
_CSV_HEADER = [
    "run_id", "experiment_id", "env_id", "algo", "seed",
    "env_steps", "wall_clock", "metric", "value",
]


def _frame_metrics(run: LoadedRun, frame: dict[str, Any]) -> list[tuple[str, float]]:
    """The ``(metric, value)`` pairs to emit for one frame — omitting any that isn't populated."""
    out: list[tuple[str, float]] = []
    reward = score_of_frame(run.config.algo, frame)
    if reward is not None:
        out.append(("reward", reward))
        sp = skill_pct(reward, run.min_score, run.solved_score)
        if sp is not None:
            out.append(("skill_pct", sp))
    for key in ("ep_len_mean", "loss"):
        val = frame.get(key)
        if val is not None:
            out.append((key, float(val)))
    return out


def build_csv(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """Tidy long CSV — one row per ``(run, frame, metric)`` at full resolution. ``pivot`` is accepted for
    a uniform registry signature but does not change the output: the tidy form carries *both* the raw and
    the normalized metric, so either pivot is a ``groupby`` in pandas."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_CSV_HEADER)
    for run in runs:
        rid, env_id, algo, seed = run.meta.id, run.config.env_id, run.config.algo, run.config.seed
        for frame in run.frames:
            env_steps = int(frame.get("env_steps", frame.get("timesteps", 0)) or 0)
            wall_clock = float(frame.get("wall_clock", frame.get("elapsed", 0.0)) or 0.0)
            for metric, value in _frame_metrics(run, frame):
                # experiment_id is a forward-compatible column (X3 fills it once runs group into
                # experiments); empty for now so consumers group by run_id.
                writer.writerow([rid, "", env_id, algo, seed, env_steps, wall_clock, metric, value])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# XLSX (publication layout)
# ---------------------------------------------------------------------------

_H1 = Font(bold=True, size=13)  # sheet / section title
_H2 = Font(bold=True)  # table header row

# The X2 summary columns, in order: (English identifier, attribute on RunSummary).
_SUMMARY_COLS: list[tuple[str, str]] = [
    ("run_id", "run_id"), ("env_id", "env_id"), ("algo", "algo"), ("seed", "seed"),
    ("n_frames", "n_frames"),
    ("final_reward", "final_reward"), ("final_skill_pct", "final_skill_pct"),
    ("solved_env_steps", "solved_env_steps"), ("solved_wall_clock", "solved_wall_clock"),
    ("auc_normalized", "auc_normalized"), ("late_reward_std", "late_reward_std"),
    ("across_seed_std", "across_seed_std"), ("mean_steps_per_sec", "mean_steps_per_sec"),
    ("final_env_steps", "final_env_steps"), ("final_wall_clock", "final_wall_clock"),
    ("peak_reward", "peak_reward"), ("peak_skill_pct", "peak_skill_pct"),
    ("collapse_pct", "collapse_pct"),
]


def _write_summary_sheet(ws: Worksheet, runs: list[LoadedRun]) -> None:
    """``Summary`` — one row per run × the X2 stats (the ± / across-seed columns fill once X4 lands)."""
    ws["A1"] = "Summary — one row per run (X2 statistics)  /  Souhrn — jeden řádek na běh"
    ws["A1"].font = _H1
    header_row = 3
    for col, (name, _) in enumerate(_SUMMARY_COLS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = _H2
    for r, run in enumerate(runs, start=header_row + 1):
        for col, (_, attr) in enumerate(_SUMMARY_COLS, start=1):
            ws.cell(row=r, column=col, value=getattr(run.summary, attr))


def _curve_metric(run: LoadedRun, frame: dict[str, Any], pivot: Pivot) -> float | None:
    """The y-value plotted per frame for the chosen pivot: raw reward (per-game) or normalized skill-%
    (per-algorithm)."""
    reward = score_of_frame(run.config.algo, frame)
    if pivot == "algo":
        return skill_pct(reward, run.min_score, run.solved_score)
    return reward


def _write_game_sheet(wb: Workbook, title: str, runs: list[LoadedRun], pivot: Pivot) -> None:
    """One per-game (or per-algorithm) sheet: a bilingual title + config header, then a downsampled curve
    table with one ``(x, y)`` column pair per run and a **native Excel scatter-line chart** over them.

    A scatter (line) chart — not a plain line chart — because the runs have *different* env-step grids
    (each trainer logs on its own cadence): a scatter series carries its own X column, so curves of
    different length/spacing overlay honestly without resampling onto a shared axis.
    """
    ws = wb.create_sheet(title=title[:31])  # Excel caps sheet names at 31 chars
    metric_label = "skill %" if pivot == "algo" else "reward"
    first = runs[0]
    ws["A1"] = f"{first.env_name_en}  /  {first.env_name_cz}"
    ws["A1"].font = _H1

    # Config header — the compared runs at a glance (label · algo · seed · budget · key hyperparameters).
    ws["A2"] = "Runs compared  /  Porovnané běhy:"
    ws["A2"].font = _H2
    row = 3
    for run in runs:
        hp = active_hyperparams(run.config)
        key_hp = ", ".join(f"{k}={v}" for k, v in list(hp.items())[:4])
        ws.cell(row=row, column=1, value=run.meta.label)
        ws.cell(row=row, column=2, value=f"{run.config.algo} · seed {run.config.seed}")
        ws.cell(row=row, column=3, value=f"budget {run.config.total_timesteps}")
        ws.cell(row=row, column=4, value=key_hp)
        row += 1

    # Curve table: a header row, then each run as an adjacent (env_steps, metric) column pair. Each run's
    # curve is LTTB-downsampled once here and its length remembered for the chart's series references.
    data_header = row + 1
    data_start = data_header + 1
    lengths: list[int] = []
    for k, run in enumerate(runs):
        xcol = 2 * k + 1
        ycol = 2 * k + 2
        xs = [int(f.get("env_steps", f.get("timesteps", 0)) or 0) for f in run.frames]
        ys = [_curve_metric(run, f, pivot) for f in run.frames]
        pts = downsample(xs, ys, _LTTB_TARGET)
        ws.cell(row=data_header, column=xcol, value=f"{run.meta.label} · env_steps").font = _H2
        ws.cell(row=data_header, column=ycol, value=f"{run.meta.label} · {metric_label}").font = _H2
        for i, (x, y) in enumerate(pts):
            ws.cell(row=data_start + i, column=xcol, value=x)
            ws.cell(row=data_start + i, column=ycol, value=y)
        lengths.append(len(pts))

    if not any(lengths):  # no plottable points across any run — skip the (empty) chart
        return

    chart = ScatterChart()
    chart.title = f"{first.env_name_en} — {metric_label} vs env_steps"
    chart.x_axis.title = "env_steps"
    chart.y_axis.title = metric_label
    chart.height = 10
    chart.width = 20
    for k, (run, n) in enumerate(zip(runs, lengths, strict=True)):
        if n == 0:
            continue
        xcol, ycol = 2 * k + 1, 2 * k + 2
        xref = Reference(ws, min_col=xcol, min_row=data_start, max_row=data_start + n - 1)
        yref = Reference(ws, min_col=ycol, min_row=data_start, max_row=data_start + n - 1)
        series = Series(yref, xref, title=str(run.meta.label))
        series.marker = Marker(symbol="none")  # a smooth line, no per-point markers
        series.graphicalProperties.line.width = 20000  # ~1.6pt — force a visible connecting line
        chart.series.append(series)
    ws.add_chart(chart, ws.cell(row=data_header, column=2 * len(runs) + 2).coordinate)


def _write_config_sheet(ws: Worksheet, runs: list[LoadedRun]) -> None:
    """``Config`` — one row per run with the standard fields + the union of active hyperparameters."""
    ws["A1"] = "Config — full hyperparameters per run  /  Konfigurace — plné hyperparametry"
    ws["A1"].font = _H1
    # Union of every run's active-hyperparameter keys (a run gets a blank where its algo lacks a key).
    hp_keys: list[str] = []
    for run in runs:
        for k in active_hyperparams(run.config):
            if k not in hp_keys:
                hp_keys.append(k)
    header = ["run_id", "label", "env_id", "algo", "seed", "total_timesteps", *hp_keys]
    for col, name in enumerate(header, start=1):
        ws.cell(row=3, column=col, value=name).font = _H2
    for r, run in enumerate(runs, start=4):
        hp = active_hyperparams(run.config)
        base = [run.meta.id, run.meta.label, run.config.env_id, run.config.algo,
                run.config.seed, run.config.total_timesteps]
        for col, val in enumerate(base, start=1):
            ws.cell(row=r, column=col, value=val)
        for col, k in enumerate(hp_keys, start=len(base) + 1):
            ws.cell(row=r, column=col, value=hp.get(k))


def _write_methods_sheet(ws: Worksheet, runs: list[LoadedRun]) -> None:
    """``Methods`` — the environment + per-run provenance + the reproducibility card. Bilingual labels."""
    ws["A1"] = "Methods & reproducibility  /  Metodika a reprodukovatelnost"
    ws["A1"].font = _H1
    facts = provenance.methods_facts()
    row = 3

    def kv(label: str, value: Any) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = _H2
        ws.cell(row=row, column=2, value=value)
        row += 1

    kv("Software / Software", "")
    for name in ("python", "torch", "stable_baselines3", "gymnasium", "numpy"):
        kv(f"  {name}", facts[name])
    kv("Hardware / Hardware", "")
    kv("  platform", facts["platform"])
    kv("  gpu", facts["gpu"])
    kv("  git_commit", facts["git_commit"])
    row += 1

    kv("Comparison axes / Srovnávací osy", "")
    kv("  env_steps", "cumulative environment interactions / kumulativní interakce s prostředím")
    kv("  wall_clock", "elapsed seconds / uplynulé sekundy")
    row += 1

    ws.cell(row=row, column=1, value="Per-run / Podle běhu").font = _H2
    row += 1
    for run in runs:
        kv(f"  {run.meta.label}",
           f"env=[{run.min_score}, {run.solved_score}] · seed={run.config.seed}")
    row += 1

    ws.cell(row=row, column=1, value="Reproducibility card / Karta reprodukovatelnosti").font = _H1
    row += 2
    for run in runs:
        card = _repro_card_text(run)
        for line in card.splitlines():
            ws.cell(row=row, column=1, value=line)
            row += 1
        row += 1


def build_xlsx(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """The publication workbook: ``Summary`` + a per-game sheet for every game with ≥1 run + ``Config`` +
    ``Methods``. ``pivot="game"`` → per-game sheets of raw reward; ``pivot="algo"`` → per-algorithm sheets
    of normalized skill-%. Curves are LTTB-downsampled; a native Excel chart overlays them."""
    wb = Workbook()
    _write_summary_sheet(wb.active, runs)
    wb.active.title = "Summary"

    # The "complete workbook": group by game (default) or by algorithm, a sheet per non-empty group.
    groups: dict[str, list[LoadedRun]] = {}
    order: list[str] = []
    for run in runs:
        key = run.config.algo if pivot == "algo" else run.config.env_id
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(run)
    used: set[str] = set()
    for key in order:
        # Distinct, ≤31-char sheet title (Excel limit); disambiguate a collision with a numeric suffix.
        base = (groups[key][0].env_name_en if pivot == "game" else key)[:28]
        title = base
        n = 1
        while title[:31] in used:
            n += 1
            title = f"{base}~{n}"
        used.add(title[:31])
        _write_game_sheet(wb, title, groups[key], pivot)

    _write_config_sheet(wb.create_sheet(title="Config"), runs)
    _write_methods_sheet(wb.create_sheet(title="Methods"), runs)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Reproducibility card (Markdown artifact)
# ---------------------------------------------------------------------------


def _repro_card_text(run: LoadedRun) -> str:
    """The reproducibility card for one run: config-hash + BibTeX + the reproduce command."""
    h = provenance.config_hash(run.config)
    bib = provenance.bibtex(run.config, created_at=run.meta.created_at, label=run.meta.label)
    cmd = provenance.reproduce_command(run.config)
    return "\n".join(
        [
            f"## {run.meta.label}",
            f"- run_id: {run.meta.id}",
            f"- env: {run.config.env_id}  ·  algo: {run.config.algo}  ·  seed: {run.config.seed}",
            f"- config-hash (sha256): {h}",
            "",
            "### BibTeX",
            bib,
            "",
            "### Reproduce",
            cmd,
        ]
    )


def build_repro_card(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """A Markdown reproducibility card for each run in the selection (config-hash + BibTeX + reproduce
    command). ``pivot`` is accepted for the uniform registry signature; the card is pivot-independent."""
    header = "# Reproducibility cards\n\nEach run's citable config-hash, BibTeX entry, and reproduce command.\n"
    body = "\n\n---\n\n".join(_repro_card_text(run) for run in runs)
    return (header + "\n" + body + "\n").encode("utf-8")


# ---------------------------------------------------------------------------
# LaTeX table (booktabs)
# ---------------------------------------------------------------------------

# (column header, RunSummary attribute, format spec) for the results table.
_LATEX_COLS: list[tuple[str, str, str]] = [
    ("Env", "env_id", "s"), ("Algo", "algo", "s"), ("Seed", "seed", "d"),
    ("Final \\%", "final_skill_pct", ".1f"), ("AUC", "auc_normalized", ".3f"),
    ("Steps-to-solve", "solved_env_steps", "d"), ("Collapse \\%", "collapse_pct", ".1f"),
]


def _latex_cell(value: Any, spec: str) -> str:
    """One formatted, LaTeX-escaped table cell; ``--`` for a missing (``None``) value."""
    if value is None:
        return "--"
    if spec == "s":
        return str(value).replace("_", "\\_")
    return format(value, spec)


def build_latex(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """A paste-ready booktabs results table of the X2 summary stats (pure text, no dependency)."""
    n = len(_LATEX_COLS)
    lines = [
        "% Requires \\usepackage{booktabs}",
        "\\begin{table}[t]",
        "  \\centering",
        "  \\caption{Training results (RL Lab export).}",
        "  \\begin{tabular}{" + "l" * 3 + "r" * (n - 3) + "}",
        "    \\toprule",
        "    " + " & ".join(h for h, _, _ in _LATEX_COLS) + " \\\\",
        "    \\midrule",
    ]
    for run in runs:
        cells = [_latex_cell(getattr(run.summary, attr), spec) for _, attr, spec in _LATEX_COLS]
        lines.append("    " + " & ".join(cells) + " \\\\")
    lines += ["    \\bottomrule", "  \\end{tabular}", "\\end{table}", ""]
    return "\n".join(lines).encode("utf-8")


# ---------------------------------------------------------------------------
# Vector figure (SVG) — a standalone, publication-ready comparison chart
# ---------------------------------------------------------------------------

# A self-contained SVG can't reach the app's CSS theme tokens, so the figure carries its own palette
# (the light-mode run-compare hues, for a white-background figure) + neutral ink. Order = selection order.
_FIG_COLORS = ["#c2820a", "#0284c7", "#e11d48", "#059669", "#7c3aed", "#4d7c0f", "#8b3fd6", "#1499c7"]
_FIG_AXIS = "#334155"
_FIG_GRID = "#e2e8f0"
_FIG_INK = "#0f172a"
_FIG_MUTED = "#64748b"
_FIG_GOAL = "#c2820a"


def _xml_escape(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _fmt_count(v: float) -> str:
    """Compact SI-ish tick label (1.2M / 500k / 320) for the env-step axis."""
    a = abs(v)
    if a >= 1e6:
        return f"{v / 1e6:.1f}M".replace(".0M", "M")
    if a >= 1e3:
        return f"{v / 1e3:.0f}k"
    return f"{v:.0f}"


def _figure_series(runs: list[LoadedRun], pivot: Pivot) -> list[tuple[str, list[tuple[float, float]]]]:
    """Each run's LTTB-downsampled ``(env_steps, metric)`` curve for the chosen pivot, dropping frames with
    no plottable y — the exact data the on-screen chart draws, re-derived server-side over full history."""
    out: list[tuple[str, list[tuple[float, float]]]] = []
    for run in runs:
        xs = [int(f.get("env_steps", f.get("timesteps", 0)) or 0) for f in run.frames]
        ys = [_curve_metric(run, f, pivot) for f in run.frames]
        pts = [(float(x), float(y)) for x, y in downsample(xs, ys, 400) if y is not None]
        out.append((run.meta.label, pts))
    return out


def build_figure(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """A standalone SVG line chart of the selected runs — a vector figure to drop straight into a paper or
    slides. Honours the pivot: ``"game"`` plots raw reward, ``"algo"`` the normalized skill-% (0–100). The
    x-axis is ``env_steps`` (the fair, hardware-independent axis). Curves are LTTB-downsampled; each run
    gets a legend entry. A goal line marks the solved score (skill 100 %, or a single game's solved reward).
    """
    W, H = 760, 460
    ml, mr, mt, mb = 66, 196, 46, 54  # legend lives in the right margin
    pw, ph = W - ml - mr, H - mt - mb
    is_skill = pivot == "algo"

    series = _figure_series(runs, pivot)
    plotted = [(label, pts) for label, pts in series if pts]

    if not plotted:
        empty = (
            f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" font-family="sans-serif">'
            f'<rect width="{W}" height="{H}" fill="#ffffff"/>'
            f'<text x="{W / 2}" y="{H / 2}" text-anchor="middle" fill="{_FIG_MUTED}" font-size="15">'
            "No plottable data in the selected runs</text></svg>"
        )
        return empty.encode("utf-8")

    xmax = max((x for _, pts in plotted for x, _ in pts), default=1.0) or 1.0
    ys_all = [y for _, pts in plotted for _, y in pts]
    ymin_data, ymax_data = min(ys_all), max(ys_all)

    # Goal line + y-range: skill is a fixed 0–100 track; raw reward auto-fits, folding in a single game's
    # solved score so the target is visible on the same axis.
    single_env = len({r.config.env_id for r in runs}) == 1
    goal: float | None = 100.0 if is_skill else (runs[0].solved_score if (single_env and runs) else None)
    if is_skill:
        ymin, ymax = 0.0, max(100.0, ymax_data)
    else:
        lo, hi = ymin_data, ymax_data
        if goal is not None:
            lo, hi = min(lo, goal), max(hi, goal)
        pad = (hi - lo) * 0.06 or 1.0
        ymin, ymax = lo - pad, hi + pad
    yspan = ymax - ymin or 1.0

    def to_x(x: float) -> float:
        return ml + (x / xmax) * pw

    def to_y(y: float) -> float:
        return mt + (1 - (y - ymin) / yspan) * ph

    parts: list[str] = [
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" '
        f'font-family="Helvetica, Arial, sans-serif">',
        f'<rect width="{W}" height="{H}" fill="#ffffff"/>',
    ]

    title = "Normalized skill (%) vs environment steps" if is_skill else "Reward vs environment steps"
    parts.append(
        f'<text x="{ml}" y="26" fill="{_FIG_INK}" font-size="16" font-weight="700">{_xml_escape(title)}</text>'
    )

    # Gridlines + ticks (5 each, linear).
    for i in range(5):
        gy = mt + (i / 4) * ph
        yval = ymax - (i / 4) * yspan
        parts.append(f'<line x1="{ml}" y1="{gy:.1f}" x2="{ml + pw}" y2="{gy:.1f}" stroke="{_FIG_GRID}" stroke-width="1"/>')
        ylab = f"{yval:.0f}%" if is_skill else _fmt_count(yval) if abs(yval) >= 100 else f"{yval:.1f}"
        parts.append(
            f'<text x="{ml - 8}" y="{gy + 4:.1f}" text-anchor="end" fill="{_FIG_MUTED}" font-size="11">{ylab}</text>'
        )
    for i in range(5):
        gx = ml + (i / 4) * pw
        xval = (i / 4) * xmax
        parts.append(
            f'<text x="{gx:.1f}" y="{mt + ph + 20:.1f}" text-anchor="middle" fill="{_FIG_MUTED}" '
            f'font-size="11">{_fmt_count(xval)}</text>'
        )

    # Goal line.
    if goal is not None and ymin <= goal <= ymax:
        gy = to_y(goal)
        parts.append(
            f'<line x1="{ml}" y1="{gy:.1f}" x2="{ml + pw}" y2="{gy:.1f}" stroke="{_FIG_GOAL}" '
            f'stroke-width="1.3" stroke-dasharray="5 4" opacity="0.85"/>'
        )
        parts.append(
            f'<text x="{ml + pw - 4}" y="{gy - 5:.1f}" text-anchor="end" fill="{_FIG_GOAL}" '
            f'font-size="10.5" font-weight="600">solved</text>'
        )

    # Axes.
    parts.append(f'<line x1="{ml}" y1="{mt}" x2="{ml}" y2="{mt + ph}" stroke="{_FIG_AXIS}" stroke-width="1.4"/>')
    parts.append(
        f'<line x1="{ml}" y1="{mt + ph}" x2="{ml + pw}" y2="{mt + ph}" stroke="{_FIG_AXIS}" stroke-width="1.4"/>'
    )
    parts.append(
        f'<text x="{ml + pw / 2:.1f}" y="{H - 12}" text-anchor="middle" fill="{_FIG_AXIS}" '
        f'font-size="12">environment steps</text>'
    )
    ylabel = "skill %" if is_skill else "reward"
    parts.append(
        f'<text transform="translate(18,{mt + ph / 2:.1f}) rotate(-90)" text-anchor="middle" '
        f'fill="{_FIG_AXIS}" font-size="12">{ylabel}</text>'
    )

    # One polyline per run + a legend entry.
    for k, (label, pts) in enumerate(plotted):
        color = _FIG_COLORS[k % len(_FIG_COLORS)]
        d = " ".join(f"{'M' if i == 0 else 'L'}{to_x(x):.1f},{to_y(y):.1f}" for i, (x, y) in enumerate(pts))
        parts.append(f'<path d="{d}" fill="none" stroke="{color}" stroke-width="1.9" stroke-linejoin="round"/>')
        ly = mt + 6 + k * 20
        lx = ml + pw + 16
        parts.append(f'<rect x="{lx}" y="{ly - 8}" width="12" height="12" rx="2" fill="{color}"/>')
        clipped = label if len(label) <= 22 else label[:21] + "…"
        parts.append(
            f'<text x="{lx + 18}" y="{ly + 2}" fill="{_FIG_INK}" font-size="11.5">{_xml_escape(clipped)}</text>'
        )

    parts.append("</svg>")
    return "".join(parts).encode("utf-8")


# ---------------------------------------------------------------------------
# TensorBoard event files — one run per log dir, zipped
# ---------------------------------------------------------------------------


def build_tensorboard(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """A ZIP of TensorBoard event files, one log directory per run — drop the unzipped folder into
    ``tensorboard --logdir`` to browse the curves interactively. Each run logs every populated metric
    (reward, normalized skill %, episode length, loss) as its own scalar tag against ``env_steps`` as the
    global step. ``pivot`` is accepted for the uniform registry signature but ignored: every metric is
    logged, so both pivots are already present. An empty selection yields a valid (empty) archive."""
    import os
    import tempfile
    import zipfile

    from torch.utils.tensorboard import SummaryWriter

    buf = io.BytesIO()
    with tempfile.TemporaryDirectory() as tmp:
        for run in runs:
            safe = f"{run.config.env_id}_{run.config.algo}_seed{run.config.seed}_{run.meta.id[:8]}"
            safe = "".join(c if c.isalnum() or c in "._-" else "_" for c in safe)
            writer = SummaryWriter(log_dir=os.path.join(tmp, safe))
            for frame in run.frames:
                step = int(frame.get("env_steps", frame.get("timesteps", 0)) or 0)
                for metric, value in _frame_metrics(run, frame):
                    writer.add_scalar(metric, value, global_step=step)
            writer.close()

        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for root, _dirs, files in os.walk(tmp):
                for name in files:
                    path = os.path.join(root, name)
                    zf.write(path, os.path.relpath(path, tmp))
    return buf.getvalue()


# ---------------------------------------------------------------------------
# NPZ score matrix (X4, Wave-2 registry addition) — the exact rliable input
# ---------------------------------------------------------------------------


def normalized_score(run: LoadedRun) -> float | None:
    """One run's scalar score for the rliable ``runs × tasks`` matrix: its final skill % (X2) mapped to a
    0–1 fraction. ``None`` when the run never produced a usable skill number (skipped from the matrix)."""
    pct = run.summary.final_skill_pct
    return None if pct is None else pct / 100.0


def build_scorematrix(runs: list[LoadedRun], pivot: Pivot = "game") -> bytes:
    """The normalized ``runs × tasks`` score matrix per algorithm, as an ``.npz`` (X4 rliable input).

    Grouped by algorithm (each a *method*): for algo ``a`` the archive holds ``a__matrix`` (rows = seeds,
    cols = tasks), ``a__tasks`` (env ids) and ``a__seeds``. Round-trips with a plain ``numpy.load`` — string
    labels use unicode dtype (no ``allow_pickle`` needed). An empty selection yields a valid empty archive.
    """
    by_algo: dict[str, list[LoadedRun]] = {}
    for run in runs:
        by_algo.setdefault(run.config.algo, []).append(run)

    arrays: dict[str, Any] = {}
    for algo, group in by_algo.items():
        entries = [(r.config.env_id, r.config.seed, normalized_score(r)) for r in group]
        sm = rliable.build_score_matrix(entries)
        arrays[f"{algo}__matrix"] = sm.matrix
        arrays[f"{algo}__tasks"] = np.asarray(sm.tasks)
        arrays[f"{algo}__seeds"] = np.asarray(sm.seeds, dtype=int)

    buf = io.BytesIO()
    np.savez(buf, **arrays)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# The format registry — a format is a thin plugin; adding one is one row here.
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExportFormat:
    """One registered output format: how to build it, its MIME type, and its file extension."""

    build: Callable[[list[LoadedRun], Pivot], bytes]
    media_type: str
    extension: str


REGISTRY: dict[str, ExportFormat] = {
    "csv": ExportFormat(build_csv, "text/csv", "csv"),
    "xlsx": ExportFormat(
        build_xlsx,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "xlsx",
    ),
    "repro": ExportFormat(build_repro_card, "text/markdown", "md"),
    "latex": ExportFormat(build_latex, "text/plain", "tex"),
    "figure": ExportFormat(build_figure, "image/svg+xml", "svg"),
    "tensorboard": ExportFormat(build_tensorboard, "application/zip", "zip"),
    "scorematrix": ExportFormat(build_scorematrix, "application/octet-stream", "npz"),
}


def export(fmt: str, run_ids: list[str], pivot: Pivot = "game") -> tuple[bytes, str, str]:
    """Run the pipeline end-to-end for one format: load the runs, dispatch to the plugin, and return
    ``(content, media_type, filename)``. Raises :class:`KeyError` for an unknown format (the caller maps
    it to a 404/400) and returns an empty payload for an empty / all-unknown run selection."""
    spec = REGISTRY[fmt]
    runs = load_runs(run_ids)
    content = spec.build(runs, pivot)
    filename = f"rl-lab-export.{spec.extension}"
    return content, spec.media_type, filename
